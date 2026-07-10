#!/usr/bin/env python3
"""Parse Google Search Console xlsx exports for the weekly Switchy SEO review.

Usage:
    parse_gsc.py [data_dir]

Finds the newest Performance-on-Search export(s) in data_dir (default:
~/Documents/2026/Get Rich/SEO/Switchy Data), parses the Queries and Pages
sheets (single-period or two-period compare exports), and prints a markdown
report with:
  - per-page metrics (with week-over-week deltas when the export compares periods)
  - title-tweak opportunities: queries at position 8-20 with impressions
  - CTR flags: queries/pages at position <=10 with impressions but ~zero clicks
Exit code 1 if the newest export is older than 7 days (stale data warning).
"""
import sys
import time
from pathlib import Path

import openpyxl

DEFAULT_DIR = Path.home() / "Documents/2026/Get Rich/SEO/Switchy Data"
STALE_DAYS = 7


def load_sheet(ws):
    rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    if not rows:
        return [], []
    header, data = rows[0], rows[1:]
    return header, [r for r in data if r and r[0] is not None]


def metric_columns(header):
    """Map metric name -> list of column indexes (2 when compare-mode, else 1).
    GSC compare headers look like '23/06/2026 - 29/06/2026 Clicks'."""
    cols = {}
    for i, h in enumerate(header[1:], start=1):
        if h is None:
            continue
        name = str(h).split()[-1].lower()  # clicks / impressions / ctr / position
        cols.setdefault(name, []).append(i)
    return cols


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def parse_performance(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    out = {}
    for sheet in ("Queries", "Pages"):
        if sheet not in wb.sheetnames:
            continue
        header, data = load_sheet(wb[sheet])
        cols = metric_columns(header)
        compare = any(len(v) > 1 for v in cols.values())
        entries = []
        for row in data:
            e = {"name": str(row[0])}
            for metric, idxs in cols.items():
                e[metric] = num(row[idxs[0]])
                if compare and len(idxs) > 1:
                    e[f"{metric}_prev"] = num(row[idxs[1]])
            entries.append(e)
        out[sheet.lower()] = {"compare": compare, "entries": entries}
    # period info, if present
    if "Filters" in wb.sheetnames:
        _, filters = load_sheet(wb["Filters"])
        out["filters"] = {str(r[0]): str(r[1]) for r in filters if len(r) > 1}
    return out


def fmt_delta(cur, prev):
    d = cur - prev
    return f"{cur:g} ({'+' if d >= 0 else ''}{d:g})"


def report(perf, fname):
    print(f"# GSC report — {fname}")
    f = perf.get("filters", {})
    if f:
        print(f"filters: {f}")
    print()

    pages = perf.get("pages", {})
    if pages:
        print("## Pages")
        for e in pages["entries"]:
            line = f"- {e['name']}: {e['impressions']:g} impressions, {e['clicks']:g} clicks, pos {e['position']:.1f}"
            if pages["compare"]:
                line += (f"  [prev: {e.get('impressions_prev', 0):g} imp, "
                         f"{e.get('clicks_prev', 0):g} clicks, pos {e.get('position_prev', 0):.1f}]")
            print(line)
        print()

    queries = perf.get("queries", {})
    if queries:
        entries = sorted(queries["entries"], key=lambda e: -e["impressions"])
        print("## All queries (by impressions)")
        for e in entries:
            print(f"- \"{e['name']}\": {e['impressions']:g} imp, {e['clicks']:g} clicks, pos {e['position']:.1f}")
        print()

        opps = [e for e in entries if 8 <= e["position"] <= 20 and e["impressions"] > 0]
        print("## Title-tweak opportunities (position 8-20)")
        if opps:
            for e in opps:
                print(f"- \"{e['name']}\": pos {e['position']:.1f}, {e['impressions']:g} impressions")
        else:
            print("- none this week")
        print()

        flags = [e for e in entries
                 if e["position"] <= 10 and e["impressions"] >= 20 and e["clicks"] == 0]
        print("## CTR flags (pos <=10, >=20 impressions, 0 clicks)")
        if flags:
            for e in flags:
                print(f"- \"{e['name']}\": pos {e['position']:.1f}, {e['impressions']:g} impressions, 0 clicks")
        else:
            print("- none this week")
        print()


def main():
    data_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_DIR
    perf_files = sorted(
        (p for p in data_dir.glob("*.xlsx") if "Performance" in p.name),
        key=lambda p: p.stat().st_mtime, reverse=True)
    if not perf_files:
        print(f"No Performance xlsx exports found in {data_dir}")
        sys.exit(1)

    newest = perf_files[0]
    age_days = (time.time() - newest.stat().st_mtime) / 86400
    print(f"Newest export: {newest.name} ({age_days:.1f} days old)")
    if age_days > STALE_DAYS:
        print(f"WARNING: newest export is older than {STALE_DAYS} days — ask the user "
              "for a fresh GSC export before drawing conclusions.")
    print()
    report(parse_performance(newest), newest.name)
    sys.exit(1 if age_days > STALE_DAYS else 0)


if __name__ == "__main__":
    main()
